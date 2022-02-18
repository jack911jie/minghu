import os
import sys
import random
from turtle import bgcolor
# from openpyxl.reader.excel import load_workbook
sys.path.append(os.path.join(os.path.dirname(os.path.abspath(__file__)),'modules'))
import pic_transfer
import days_cal
import readconfig
import composing
import get_data
import draw_pic
import math
import json
import openpyxl
# from openpyxl import load_workbook
import pandas as pd
pd.set_option('display.unicode.ambiguous_as_wide', True)
pd.set_option('display.unicode.east_asian_width', True)
import numpy as np
from datetime import datetime
import re
import random
from PIL import Image,ImageDraw,ImageFont
import random
import matplotlib.pyplot as plt 
import matplotlib.font_manager as fm
import matplotlib.ticker as mticker
from tkinter import simpledialog

# from matplotlib.backends.backend_agg import FigureCanvasAgg
plt.rcParams['font.sans-serif']=['SimHei']  # 黑体

class MingHu:
    def __init__(self,place='minghu',adj_bfr='yes',adj_src='prg',gui=''):
        self.dir=os.path.dirname(os.path.abspath(__file__))
        config=readconfig.exp_json(os.path.join(self.dir,'configs','main_'+place+'.config'))
        self.cus_file_dir=config['会员档案文件夹']
        self.material_dir=config['素材文件夹']
        self.ins_dir=config['教练文件夹']
        self.slogan_dir=config['文案文件夹']
        self.save_dir=config['输出文件夹']
        self.public_dir=config['公共素材文件夹']
        self.adj_bfr=adj_bfr
        self.adj_src=adj_src
        self.gui=gui
        self.place=place
        print(os.path.join(self.dir,'configs','main_'+place+'.config'),self.ins_dir,self.cus_file_dir)
        self.df_ins=pd.read_excel(os.path.join(self.ins_dir,'教练信息.xlsx'),sheet_name='教练信息')
        self.color_config_fn=os.path.join(os.path.dirname(__file__),'configs','colors.config')
        with open(os.path.join(self.material_dir,'txt_public.txt'),'r',encoding='utf-8') as txt_pub:
            self.txt_public=txt_pub.readlines()
        self.cus_instance_name=self.txt_public[5].strip()[0:2]
        self.prefix=self.cus_instance_name[0:2]
        self.gym_name=self.txt_public[4]
        self.gym_addr=self.txt_public[3]
        if '%' in self.gym_addr:
            self.gym_addr=''
        self.txt_ins_word=self.txt_public[2]
        self.txt_mini_title=self.txt_public[1]
        self.txt_slogan=self.txt_public[0]

    def auto_cus_xls(self,cus_name_input='',mode='prgrm',gui=''):
        # cus_name_input=''
        if mode=='prgrm':
            while cus_name_input=='':
                cus_name_input=input('请输入新会员姓名：')
                if cus_name_input=='exit':
                    exit(0)
        elif mode=='gui':
            cus_name_input=cus_name_input
 
        nums=[]
        for fn in os.listdir(self.cus_file_dir):
            if len(fn)<16:
                if re.match(self.prefix+r'\d\d\d.*.xlsx',fn):
                    num=int(fn[2:5])
                    if num not in nums:
                        nums.append(num)

        new_num=str(max(nums)+1).zfill(3)
        if mode=='prgrm':
            verify=input('\n新会员档案文件编号为：{}，确认直接按回车。\n如需自行修改编号，请输入编号后再回车。\n请选择——————'.format(self.prefix+new_num+cus_name_input))
        elif mode=='gui':
            gui.delete('1.0','end')
            print('\n新会员档案文件编号为：{}，确认直接按回车。\n如需自行修改编号，请输入编号后再回车。\n请选择——————'.format(self.prefix+new_num+cus_name_input))
            # verify=''
            while True:
                verify = simpledialog.askstring(title="是否修改编号？",prompt="请输入新编号（三位数字）")
                if not verify:
                    break
                else:
                    if len(verify)==3 and re.match(r'\d\d\d',verify):
                        break                        
                    else:
                        gui.delete('1.0','end')
                        print('编号格式错误，请输入三位数字。')          
            gui.delete('1.0','end')
        if verify:
            xls_name=self.prefix+verify+cus_name_input
        else:
            xls_name=self.prefix+new_num+cus_name_input
        
        wb=openpyxl.load_workbook(os.path.join(os.path.dirname(self.cus_file_dir),'模板.xlsx'))
        sht=wb['基本情况']
        sht['A2']=xls_name[0:5]
        sht['B2']=cus_name_input
        if len(cus_name_input)>1:
            sht['C2']=cus_name_input[1:]
        else:
            sht['C2']=cus_name_input
        
        wb.save(os.path.join(self.cus_file_dir,xls_name+'.xlsx'))
        print('\n生成新的会员档案文件：{}'.format(self.cus_file_dir+'\\'+xls_name+'.xlsx'))

        return xls_name


    def fonts(self,font_name,font_size):
        fontList=readconfig.exp_json(os.path.join(self.dir,'configs','FontList.minghu.config'))
        # print(fontList)
        return ImageFont.truetype(fontList[font_name],font_size)

    def color_list(self,sex='美女',color_name=''):
        color_config=readconfig.exp_json(self.color_config_fn)
        # print(color_config['light_pink'])

        if sex=='美女':
            if color_name=='':
                color_name='light_pink'           
        elif sex=='帅哥':
            if color_name=='':
                color_name='strong_blue'
        else:
            pass
        colors=color_config['Summary'][color_name]

        return colors

    def put_txt_img(self,img,t,total_dis,xy,dis_line,fill,font_name,font_size,addSPC='None'):
        
        fontInput=self.fonts(font_name,font_size)            
        if addSPC=='add_2spaces': 
            ind='yes'
        else:
            ind='no'
            
        # txt=self.split_txt(total_dis,font_size,t,Indent='no')
        txt,p_num=composing.split_txt_Chn_eng(total_dis,font_size,t,Indent=ind)

        # font_sig = self.fonts('丁永康硬笔楷书',40)
        draw=ImageDraw.Draw(img)   
        # logging.info(txt)
        n=0
        for t in txt:              
            m=0
            for tt in t:                  
                x,y=xy[0],xy[1]+(font_size+dis_line)*n
                if addSPC=='add_2spaces':   #首行缩进
                    if m==0:    
                        # tt='  '+tt #首先前面加上两个空格
                        # logging.info('字数：'+str(len(tt))+'，坐标：'+str(x)+','+str(y))
                        # logging.info(tt)
                        draw.text((x-font_size*0.2,y), tt, fill = fill,font=fontInput) 
                    else:                       
                        # logging.info('字数：'+str(len(tt))+'，坐标：'+str(x)+','+str(y))
                        # logging.info(tt)
                        draw.text((x,y), tt, fill = fill,font=fontInput)  
                else:
                    # logging.info('字数：'+str(len(tt))+'，坐标：'+str(x)+','+str(y))
                    # logging.info(tt)
                    draw.text((x,y), tt, fill = fill,font=fontInput)  
 
                m+=1
                n+=1

    def draw(self,cus='MH001韦美霜',ins='MHINS002韦越棋',start_time='20150101',end_time=''):
        
        def slogan():
            df_slogans=pd.read_excel(os.path.join(self.slogan_dir,'文案.xlsx'))
            # print(df_slogans)
            df_slogans.dropna(axis=0,how='any',subset=['文案'],inplace=True)
            slogans=df_slogans['文案'].values.tolist()
            slogan=random.choice(slogans)
            # print('189 line:',df_slogans)
            return slogan

        def ins_info():
            df_ins=pd.read_excel(os.path.join(self.ins_dir,'教练信息.xlsx'))
            ins_inf={}
            ins_inf['nickname']=df_ins[df_ins['员工编号']==ins[0:8]]['昵称'].values.tolist()[0].strip()
            tel=str(df_ins[df_ins['员工编号']==ins[0:8]]['电话'].values.tolist()[0]).strip()
            tel=tel[0:3]+'-'+tel[3:7]+'-'+tel[7:]
            ins_inf['tel']='电话：'+tel
            return ins_inf

        def txts():
            cus_data=get_data.ReadAndExportData(adj_bfr=self.adj_bfr,adj_src=self.adj_src,gui=self.gui)
            infos=cus_data.exp_cus_prd(cus_file_dir=self.cus_file_dir,cus=cus,start_time=start_time,end_time=end_time)    
                
            # print(infos) 
            txts=Vividict()
            #文字
            txts['nickname']=infos['nickname']
            sex=infos['sex']
            if sex=='女':
                sex='美女'
            elif sex=='男':
                sex='帅哥'
            else:
                pass

            txts['sex']=sex
            txts['age']=infos['age']
        
            #测量
            if infos['body']:
                latest_msr_time=infos['body']['time']
                txts['latest_msr_time']=datetime.strftime(latest_msr_time,'%Y年%m月%d日')
                txts['ht']='身高 '+str('{:g}'.format(infos['body']['ht']))+' cm'
                txts['wt']='体重 '+str('{:g}'.format(infos['body']['wt']))+' Kg'
                txts['bfr']='体脂率 '+str(infos['body']['bfr'])
                txts['chest']='胸围 '+str('{:g}'.format(infos['body']['chest']))+' cm'
                txts['l_arm']='左臂围 '+str('{:g}'.format(infos['body']['l_arm']))  +' cm'
                txts['r_arm']='右臂围 '+str('{:g}'.format(infos['body']['r_arm'])) +' cm'
                txts['waist']='腰围 '+str('{:g}'.format(infos['body']['waist'])) +' cm'
                txts['hip']='臀围 '+str('{:g}'.format(infos['body']['hip']))  +' cm'
                txts['l_leg']='左大腿围 '+str('{:g}'.format(infos['body']['l_leg']))  +' cm'
                txts['r_leg']='右大腿围 '+str('{:g}'.format(infos['body']['r_leg']))  +' cm'
                txts['l_calf']='左小腿围 '+str('{:g}'.format(infos['body']['l_calf']))  +' cm'
                txts['r_calf']='右小腿围 '+str('{:g}'.format(infos['body']['r_calf']))  +' cm'
                txts['bfr']='体脂率：'+str('{:.2%}'.format(infos['body']['bfr']))
            else:
                txts['latest_msr_time']=0

            #训练情况
            # print('242line:',infos['train'])
            if infos['train']:
                t_muscle=''
                t_oxy=''
                for items in infos['train']:
                    # txts['train_content'][items]='    '+str(infos[items])+' 次'
                    # print(items)
                    if items=='muscle':
                        if infos['train'][items]:
                            for k in infos['train'][items]:
                                t_muscle=t_muscle+str(k)+'    '+str(infos['train'][items][k])+'次\n'
                        else:
                            t_muscle=''
                    elif items=='oxy_time':
                        _oxy_time=infos['train'][items]
                        if _oxy_time!=0:
                            if _oxy_time>60:
                                if _oxy_time%60==0:
                                    _oxy_time='有氧训练    '+str(int(_oxy_time//60))+'分钟\n'
                                else:
                                    _oxy_time='有氧训练    '+str(int(_oxy_time//60))+'分钟'
                                    # _oxy_time='有氧训练    '+str(int(_oxy_time//60))+'分钟'+str(int(_oxy_time%60))+'秒\n'
                            t_oxy=t_oxy+_oxy_time
                            t_oxy.rstrip()
                        else:
                            t_oxy=''

                t=t_muscle+t_oxy
                txts['train_content']=t.rstrip()
            else:
                txts['train_content']=''
            
    

            if txts['train_content']:
                intervals_input=infos['interval_input'][1]-infos['interval_input'][0]
                if intervals_input.days==0:
                    txts['intervals_train_0']='今天的你非常棒，因为，'
                    txts['intervals_train_1']=''
                else:
                    txts['intervals_train_0']='您在{0}-{1}的'.format(datetime.strftime(infos['interval_input'][0],'%Y年%m月%d日'),datetime.strftime(infos['interval_input'][1],'%Y年%m月%d日'))
                    txts['intervals_train_1']='{0}天里锻炼了{1}次'.format(str(intervals_input.days+1),str(infos['train_times']))
            else:
                txts['intervals_train_0']=''
                txts['intervals_train_1']=''

            # print(txts)

            radar_data={'ht_lung':infos['body']['ht_lung'],'balance':infos['body']['balance'],'power':infos['body']['power'], \
                        'flexibility':infos['body']['flexibility'],'core':infos['body']['core']}
        
            return {'txts':txts,'radar_data':radar_data}

        def radar(data,sex):
            color=self.color_list(sex=sex)
            # print(data)
            # 构造数据
            values = list(data.values())
            feature =list(data.keys()) 

            N = len(values)
            # 设置雷达图的角度，用于平分切开一个圆面
            angles = np.linspace(0, 2 * np.pi, N, endpoint=False)


            # 为了使雷达图一圈封闭起来，需要下面的步骤
            values = np.concatenate((values, [values[0]]))
            angles = np.concatenate((angles, [angles[0]]))

            # print(values,angles)

            # 绘图
            fig = plt.figure(figsize=(6,5))
            # 这里一定要设置为极坐标格式
            ax = fig.add_subplot(111, polar=True)
            # ccl=ax.patch

            # 绘制折线图
            ax.plot(angles, values, 'o-', linewidth=2,color=color['txt_train'])
            # 填充颜色
            ax.fill(angles, values, color=color['txt_train'],alpha=0.25)
            # 添加每个特征的标签
            ax.set_thetagrids(angles * 180 / np.pi, '',color='r',fontsize=13)
            # 设置雷达图的范围
            r_distance=10
            ax.set_rlim(0, r_distance)

            ax.grid(color='#F1E0D6', alpha=0.25, lw=3)
            ax.spines['polar'].set_color('#F1E0D6')
            ax.spines['polar'].set_alpha(0.2)
            ax.spines['polar'].set_linewidth(2)
            # ax.spines['polar'].set_linestyle('-.')

            #项目名称：
            a=[0,0,np.pi/30,-np.pi/50,0,0,0]
            b=[r_distance*1.1,r_distance*1.1,r_distance*1.3,r_distance*1.4,r_distance*1.12]

            e_to_c={'ht_lung': '心肺', 'balance': '平衡', 'power': '力量', 'flexibility': '柔韧性', 'core': '核心'}
            for k,i in enumerate(angles):
                try:
                    # print(k,i,e_to_c[feature[k]])
                    ax.text(i+a[k],b[k],e_to_c[feature[k]],fontsize=18,color=color['txt_train'])
                except:
                    pass

            #分值：
            # c = [1, 0.6, 1.6, 2.3, 1.5, 1,1]
            # print(len(angles))
            # for j,i in enumerate(angles):
            #     try:
            #         r=values[j]-2*i/np.pi
            #         ax.text(i,values[j]+c[j],values[j],color='#218FBD',fontsize=18)
            #     except:
            #         pass

            # 添加标题
            #plt.title('活动前后员工状态表现')
            # 添加网格线
            ax.grid(True,color='grey',alpha=0.1)

            # a=np.arange(0,2*np.pi,0.01)
            # ax.plot(a,10*np.ones_like(a),linewidth=2,color='b')


            ax.set_yticklabels([])
            # plt.savefig(savefilename,transparent=True,bbox_inches='tight')
            # 显示图形
            # plt.show()

            #将matplotlib的图形转换为PIL的对象
            image=pic_transfer.mat_to_pil_img(fig)


            return image

        def save_pic_name(cus):
            save_dir=os.path.join(self.save_dir,cus)
            if not os.path.exists(save_dir):
                os.makedirs(save_dir)
            _date=datetime.strftime(datetime.now(),"%Y%m%d_%H%M%S")
            save_name=os.path.join(save_dir,_date+'_'+cus+'.jpg')
            print('文件名：'+save_name)
            os.startfile(save_dir)
            return save_name

        def exp_pic(dat):           

            t=dat['txts']
            color=self.color_list(sex=t['sex'])
            radar_data=dat['radar_data']
            
            # print('279 line:',t)
            dis_line=20
            ft_size=36
            num_prgr=len(t['train_content'].split('\n'))
            block_wid=684

            s_top=40
            gap=20
            s_name=180
            s_title=40
            # r2=15
            if t['latest_msr_time']==0:
                s_title_body=0
                s_body=0
                gap_body=0
            else:
                s_title_body=s_title
                s_body=960
                gap_body=gap

            if t['train_content']:
                s_train_content=dis_line*(num_prgr-1)+ft_size*num_prgr+60
                if s_train_content<300:
                    s_train_content=300
                s_train=s_train_content+300
                gap_train=gap
                s_title_train=s_title
            else:
                s_title_train=0
                s_train_content=0
                s_train=0
                gap_train=0

            slogan_txt=slogan()
            slogan_txt=slogan_txt+'\n'+self.txt_public[2]
            dis_line_slogan=15
            ft_size_slogan=36
            # print('327 line',slogan_txt)
            num_prgr_slogan=len(slogan_txt.split('\n'))
            # s_slogan=120
            s_slogan=dis_line_slogan*(num_prgr_slogan-1)+ft_size_slogan*num_prgr_slogan+40
            # print(s_slogan,num_prgr_slogan)

            s_logo=680
            s_bottom=40

            y0=0
            y_name=y0+s_top+gap

            y_title_body=y_name+s_name+gap
            y_body=y_title_body+s_title_body

            y_title_train=y_body+s_body+gap_body
            y_train=y_title_train+s_title_train

            y_slogan=y_train+s_train+gap_train
            y_logo=y_slogan+s_slogan+gap
            y_bottom=y_logo+s_logo+gap

            x_l=18
            x_r=x_l+block_wid

            def bg(ins=ins):           
                # y_item=dis_line*(num_prgr-1)+ft_size*num_prgr+50
                img = Image.new("RGB",(720,y_bottom+s_bottom),(255,255,255))
                
                draw=ImageDraw.Draw(img)

                #--------框-----------
                draw.rectangle((0,0,720,y0+s_top),fill=color['comment_bg']) #top
                # draw.rectangle((x_l,y_name,x_r,y_name+s_name),fill='#fff4ee') #name
                y_pic_box=y_name+int(s_name*0.2/2)
                draw.rectangle((x_l+20,y_pic_box,x_l+20+int(s_name*0.8),y_name+int(s_name*0.9)),fill=color['train_content_bg']) #head pic box

                if t['latest_msr_time']!=0:
                    draw.rectangle((x_l,y_title_body,x_l+254,y_title_body+s_title_body),fill=color['title_bg']) #body title
                    draw.rectangle((x_l,y_body,x_r,y_body+s_body),fill=color['comment_bg']) #body

                if t['train_content']:
                    draw.rectangle((x_l,y_title_train,x_l+254,y_title_train+s_title_train),fill=color['title_bg']) #train title
                    draw.rectangle((x_l,y_train,x_r,y_train+s_train),fill=color['comment_bg']) #train
                    y_train_content_bottom=y_train+200+s_train_content
                    draw.rectangle((x_l+40,y_train+200,x_r-40,y_train_content_bottom),fill=color['train_content_bg']) #train content                    
                draw.rectangle((x_l,y_slogan,x_r,y_slogan+s_slogan),fill=color['logo_bg']) #slogan
                draw.rectangle((x_l,y_logo,x_r,y_logo+s_logo),fill=color['logo_bg']) #logo
                draw.rectangle((0,y_bottom,720,y_bottom+s_bottom),fill=color['logo_bg']) #bottom

                 #--------图片-----------

                #头像
                if t['sex']=='美女':
                    pics_f=[]
                    for fn in os.listdir(self.material_dir):
                        if re.match(r'女性头像\d{2}.jpg',fn) or re.match(r'女性头像\d{2}.png',fn):
                            pics_f.append(fn)
                    filename=random.choice(pics_f)                   
                    pic_head_src=os.path.join(self.material_dir,filename)
                else:
                    pics_f=[]
                    for fn in os.listdir(self.material_dir):
                        if re.match(r'男性头像\d{2}.jpg',fn) or re.match(r'男性头像\d{2}.png',fn):
                            pics_f.append(fn)
                    filename=random.choice(pics_f)                   
                    pic_head_src=os.path.join(self.material_dir,filename)
                    # pic_head_src=os.path.join(self.material_dir,'男性头像03.png')
                    # pass #男性
                    # pass
                pic_head=Image.open(pic_head_src)
                pic_head=pic_transfer.round_corner(pic_head)
                w_head,h_head=pic_head.size
                pic_head=pic_head.resize((int(w_head*120/h_head),120))
                r1,g1,b1,a1=pic_head.split()
                img.paste(pic_head,(x_l+20+int((s_name*0.8-pic_head.size[0])/2),y_name+30),mask=a1)

                #模特
                if t['latest_msr_time']!=0:
                    if t['sex']=='美女':
                        model_pic='size_model_female.png'
                    elif t['sex']=='帅哥':
                        model_pic='size_model_male.png'
                    else:
                        pass
                    model_src=os.path.join(self.material_dir,model_pic)
                    pic_model=Image.open(model_src)
                    w_model,h_model=pic_model.size
                    pic_model=pic_model.resize((280,int(h_model*280/w_model)))
                    r2,g2,b2,a2=pic_model.split()
                    img.paste(pic_model,(x_l+int((block_wid-pic_model.size[0])/2),y_title_body+175),mask=a2)

                    #雷达图
                    img_radar=radar(radar_data,sex=t['sex'])
                    # img_radar.show()
                    # print(img_radar.size)
                    img_radar=img_radar.resize((400,int(400*img_radar.size[1]/img_radar.size[0])))
                    img.paste(img_radar,(x_l+int((block_wid-img_radar.size[0])/2),y_title_body+175+int(h_model*280/w_model)+100))

                if t['train_content']:
                    teach_pic_src=os.path.join(self.material_dir,'指导.png')
                    pic_teach=Image.open(teach_pic_src)
                    w_teach,h_teach=pic_teach.size
                    pic_teach=pic_teach.resize((150,int(h_teach*150/w_teach)))
                    r3,g3,b3,a3=pic_teach.split()
                    img.paste(pic_teach,(x_r-150-40,y_train+200+s_train_content-150),mask=a3)
                    # img.paste(pic_teach,())   x_l+40,y_train+200,x_r-40,y_train+200+s_train_content

                #logo
                logo=Image.open(os.path.join(self.public_dir,'logo及二维码','logo.png'))
                w_logo,h_logo=logo.size
                logo=logo.resize((300,int(h_logo*300/w_logo)))
                r4,g4,b4,a4=logo.split()
                img.paste(logo,(int(x_l+(s_logo-300)/2),y_logo+30),mask=a4)

                #qrcode

                qrcode=Image.open(os.path.join(self.ins_dir,ins+'二维码.jpg'))
                w_qrcode,h_qrcode=qrcode.size
                qrcode=qrcode.resize((150,int(h_qrcode*150/w_qrcode)))
                # r5,g5,b5,a5=qrcode.split()
                img.paste(qrcode,(int(x_l+(s_logo-150)/2),y_logo+logo.size[1]+220))

                #------文字-----------
                if t['sex']=='美女':
                    title_01='看看棒棒的自己'
                elif t['sex']=='帅哥':
                    title_01='看看很酷的自己'
                else:
                    pass

                x_nickname=250
                draw.text((x_nickname,110), t['nickname'], fill = color['txt_person'],font=self.fonts('汉仪心海楷体w',80))  #姓名
                if t['sex']=='帅哥':
                    sex='先生'
                elif t['sex']=='美女':
                    sex='女士'
                else:
                    print('warning:性别有误')
                draw.text((x_nickname+len(t['nickname'])*80+30,150), sex, fill = color['txt_person'],font=self.fonts('汉仪心海楷体w',40))  #性别
                if t['latest_msr_time']!=0:
                    draw.text((x_l+30,y_title_body+5), title_01, fill = color['txt_title'],font=self.fonts('上首金牛',30))  #看看棒棒的自己
                    draw.text((x_l+115,y_title_body+65), '您最近一次测量身体围度，是在', fill = color['txt_fix'],font=self.fonts('aa楷体',36))  #您最近一次测量身体围度
                    draw.text((x_l+205,y_title_body+115), t['latest_msr_time'], fill = color['txt_date'],font=self.fonts('aa楷体',40))  #测围度日期

                    draw.text((x_l+20,y_title_body+190), t['r_arm'], fill = color['txt_dimension'],font=self.fonts('杨任东石竹体',25))  #右臂
                    draw.text((x_l+75,y_title_body+270), t['hip'], fill = color['txt_dimension'],font=self.fonts('杨任东石竹体',25))  # 臀
                    draw.text((x_l+20,y_title_body+380), t['r_leg'], fill = color['txt_dimension'],font=self.fonts('杨任东石竹体',25))  #右大腿
                    draw.text((x_l+20,y_title_body+460), t['r_calf'],fill = color['txt_dimension'],font=self.fonts('杨任东石竹体',25))  #右小腿
                    draw.text((x_l+500,y_title_body+190), t['chest'], fill = color['txt_dimension'],font=self.fonts('杨任东石竹体',25))  #胸
                    draw.text((x_l+500,y_title_body+240), t['l_arm'], fill = color['txt_dimension'],font=self.fonts('杨任东石竹体',25))  #左臂
                    draw.text((x_l+500,y_title_body+280), t['waist'], fill = color['txt_dimension'],font=self.fonts('杨任东石竹体',25))  #腰
                    draw.text((x_l+500,y_title_body+370), t['l_leg'], fill = color['txt_dimension'],font=self.fonts('杨任东石竹体',25))  #左大腿
                    draw.text((x_l+500,y_title_body+470), t['l_calf'], fill = color['txt_dimension'],font=self.fonts('杨任东石竹体',25))  #左小腿
                    draw.text((x_l+180,y_title_body+550), t['wt'], fill = color['txt_dimension'],font=self.fonts('aa楷体',25))  #体重
                    draw.text((x_l+360,y_title_body+550), t['bfr'], fill = color['txt_dimension'],font=self.fonts('aa楷体',25))  #体脂率         

                if t['train_content']:
                    draw.text((x_l+30,y_title_train+5), '看看努力的自己', fill = color['txt_title'],font=self.fonts('上首金牛',30))  #看看努力的自己                    
                    if t['intervals_train_1']:
                        draw.text((x_l+45,y_train+35), t['intervals_train_0'], fill =color['txt_fix'],font=self.fonts('aa楷体',34))  #您在。。。
                        ft_size_days=40
                        x_days=composing.center_align_x(start_x=x_l,wide=block_wid,ft_size=ft_size_days,t=t['intervals_train_1'])
                        draw.text((x_days,y_train+85), t['intervals_train_1'], fill =color['txt_train'],font=self.fonts('aa楷体',ft_size_days))  #XX天里（居中）
                        draw.text((x_l+180,y_train+140), '完成了下面的训练内容', fill = color['txt_fix'],font=self.fonts('aa楷体',32))  #完成了下面的训练内容
                        self.put_txt_img(img,t=t['train_content'],total_dis=420,xy=[x_l+95,y_train+230],dis_line=16,fill=color['txt_train'],font_name='杨任东石竹体',font_size=38)
                        if self.place=='minghu':
                            percent=random.randint(70,93)
                            draw.text((x_l+145,y_train_content_bottom+20), '击败了铭湖健身 {} 的会员!'.format(str(percent)+'%'), fill = color['txt_train'],font=self.fonts('aa楷体',32))  #击败了
                    else:
                        draw.text((x_l+145,y_train+45), t['intervals_train_0'], fill = color['txt_fix'],font=self.fonts('aa楷体',40))  #您在。。。
                        draw.text((x_l+160,y_train+85), t['intervals_train_1'], fill = color['txt_train'],font=self.fonts('aa楷体',40))  #XX天里
                        draw.text((x_l+50,y_train+115), '你完成了下面的训练内容:', fill = color['txt_fix'],font=self.fonts('aa楷体',40))  #完成了下面的训练内容
                        self.put_txt_img(img,t=t['train_content'],total_dis=420,xy=[x_l+95,y_train+230],dis_line=16,fill=color['txt_train'],font_name='杨任东石竹体',font_size=36)
                        draw.text((x_l+55,y_train_content_bottom+20), '保持这样的状态，好身材还远吗？', fill =color['txt_train'],font=self.fonts('aa楷体',40))  #击败了

                # 鸡汤
                draw.text((x_l+20,y_slogan+15),slogan_txt,fill=color['txt_slogan'],font=self.fonts('优设标题黑',ft_size_slogan))

                # addr
                #地址
                x_add=x_l+(block_wid-composing.char_len(self.txt_public[3])*30)//2
                draw.text((x_add,y_logo+240),self.gym_addr,fill=color['gym_info'],font=self.fonts('微软雅黑',30))
                #slogan
                draw.text((x_l+125,y_logo+310),self.txt_slogan,fill=color['gym_info'],font=self.fonts('丁永康硬笔楷书',60))

                ins=ins_info()
                draw.text((x_l+255,y_logo+570),ins['nickname'],fill=color['gym_info'],font=self.fonts('丁永康硬笔楷书',50))
                draw.text((x_l+115,y_logo+630),ins['tel'],fill=color['gym_info'],font=self.fonts('丁永康硬笔楷书',40))

                save_name=save_pic_name(cus)
                print(cus)
                img.save(save_name,quality=95,subsampling=0)
                print('完成')

                # img.show()

            bg()

        t=txts()
        # radar(t['radar_data'])
        exp_pic(t)
        # slogan()
        # ins_info()

class GroupDataInput:
    def __init__(self,place):
        self.dir=os.path.dirname(os.path.abspath(__file__))
        config=readconfig.exp_json(os.path.join(self.dir,'configs','main_'+place+'.config'))
        self.grp_dir=config['会员档案文件夹']

    def data_input(self):
        grp_file=os.path.join(self.grp_dir,'00-团课分班录入表.xlsx')
        df_grp_pre=pd.read_excel(grp_file,sheet_name='分组',skiprows=1)
        df_grp=df_grp_pre.iloc[:,4:]
        df_grp_names=df_grp.columns.tolist()
        #需录入数据的名单
        df_real_list=df_grp_pre['Unnamed: 0'].dropna()           

        df_data=pd.read_excel(grp_file,sheet_name='训练情况')
        df_data=df_data.iloc[1:]
        # wb_train_data=openpyxl.load_workbook(grp_file)
        # sht_train=wb_train_data['训练情况']
        # print(sht_train['i21'])

        if df_real_list.empty:
            exit('未录入数据')
        else:
            cus_list=df_real_list.apply(lambda x:x+'.xlsx').tolist()
        
        # cus_list=['MH024刘婵桢.xlsx']
        for cus_name in cus_list:
            fn=os.path.join(self.grp_dir,cus_name)

            print('正在写入{}'.format(cus_name[0:-5]))

            book=openpyxl.load_workbook(fn)
            df_to_write=pd.read_excel(fn,sheet_name='训练情况')
            # df_new=pd.concat([df_to_write,df_data])
            # print(df_new)
            writer = pd.ExcelWriter(fn,engine='openpyxl')#可以向不同的sheet写入数据      
            writer.book=book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            df_rows = df_to_write.dropna(axis=0,how='all').shape[0] #去除所有为na的行，然后获取原数据的行数
            df_data.to_excel(writer, sheet_name='训练情况',startrow=df_rows+1, index=False, header=False)#将数据写入excel中的aa表,从第一个空行开始写
            writer.save()#保存
            writer.close()

            wb=openpyxl.load_workbook(fn)
            sht=wb['训练情况']
            cols_d_e=sht['D:E']
            cols_f_h=sht['F:H']

            #日期格式
            cols_a=sht['A']
            for cell in sht['A']:
                # for cell in cells:                            
                cell.number_format= 'YYYY-MM-DD'

            # 单元格填充背景色
            bg_color_blue = openpyxl.styles.PatternFill(fgColor='DAEEF3', fill_type='solid')
            bg_color_orange = openpyxl.styles.PatternFill(fgColor='FDE9D9', fill_type='solid')
            for cells in cols_d_e:
                for cell in cells:
                    cell.fill=bg_color_blue
            
            for cells in cols_f_h:
                for cell in cells:
                    cell.fill=bg_color_orange

            
            
            wb.save(fn)

        print('完成')

class FeedBackAfterClass:
    def __init__(self,place='minghu'):
        self.dir=os.path.dirname(os.path.abspath(__file__))
        config=readconfig.exp_json(os.path.join(self.dir,'configs','main_'+place+'.config'))
        self.cus_file_dir=config['会员档案文件夹']
        self.material_dir=config['素材文件夹']
        self.ins_dir=config['教练文件夹']
        self.slogan_dir=config['文案文件夹']
        self.save_dir=config['输出文件夹']
        self.public_dir=config['公共素材文件夹']
        self.exp_knlg_dir=config['专业资料文件夹']
        self.save_dir_feedback=config['课后反馈文件夹']
        self.font_config=os.path.join(self.dir,'configs','fontList.minghu')
        self.df_ins=pd.read_excel(os.path.join(self.ins_dir,'教练信息.xlsx'),sheet_name='教练信息')
        self.color_config_fn=os.path.join(os.path.dirname(__file__),'configs','colors.config')
        with open(os.path.join(self.material_dir,'txt_public.txt'),'r',encoding='utf-8') as txt_pub:
            self.txt_public=txt_pub.readlines()
        self.cus_instance_name=self.txt_public[5].strip()[0:2]
        self.prefix=self.cus_instance_name[0:2]
        self.gym_name=self.txt_public[4]
        self.gym_addr=self.txt_public[3]
        self.txt_ins_word=self.txt_public[2]
        self.txt_mini_title=self.txt_public[1]
        self.txt_slogan=self.txt_public[0]


    def export(self,cus='MH024刘婵桢',ins='MHINS002韦越棋',date_input='20210324'):
        start_time=date_input
        end_time=date_input
        cus_data=get_data.ReadAndExportDataNew(adj_bfr='no',)
        data=cus_data.exp_cus_prd(cus_file_dir=self.cus_file_dir,cus=cus,start_time=start_time,end_time=end_time)
        return data

    def icon(self,ico,ico_size=(20,20)):
        ico=ico.resize(ico_size)
        return (ico,ico.split()[3])

    def draw(self,cus='MH024刘婵桢',ins='MHINS002韦越棋',date_input='20210324',open_dir='yes'):
        #公共文字
        # with open(os.path.join(self.material_dir,'txt_public.txt'),'r',encoding='utf-8') as txt_pub:
        #     txt_public=txt_pub.readlines()

        #文字内容
        data=self.export(cus=cus,ins=ins,date_input=date_input)
        # print(data)
        #日期
        txt_date=date_input[:4]+'年'+date_input[4:6]+'月'+date_input[6:]+'日'
        #姓名
        nickname=data['nickname']
        #性别
        sex=data['sex']
        if sex=='女':
            sex='女士'
        elif sex=='男':
            sex='先生'
        else:
            sex=''
        
        #标题框文字
        # txt_title_box='看看今天你的汗水洒在哪里？'
        txt_title_box=self.txt_mini_title
        

        #抗阻内容
        txt_train_muscle=''
        for mscl_item in data['train']['muscle_item']:
            if mscl_item[2]>0:
                txt_train_muscle=txt_train_muscle+mscl_item[0]+'  '+str(int(mscl_item[2]))+'个'+'\n'
            else:
                txt_train_muscle=txt_train_muscle+mscl_item[0]+'  '+str(int(mscl_item[3]))+'米'+'\n'
        txt_train_muscle.strip()

        #有氧内容
        txt_train_oxy=''
        for oxy_item in  data['train']['oxy_infos']:
            if oxy_item[1]%60==0:
                txt_train_oxy=txt_train_oxy+oxy_item[0]+'  '+str(int(oxy_item[1]//60))+'分'+'\n'
            else:
                # txt_train_oxy=txt_train_oxy+oxy_item[0]+'  '+str(int(oxy_item[1]//60))+'分'+'\n'
                txt_train_oxy=txt_train_oxy+oxy_item[0]+'  '+str(int(oxy_item[1]//60))+'分'+str(int(oxy_item[1]%60))+'秒\n'
        txt_train_oxy.strip()


        txt_train=txt_train_muscle+txt_train_oxy

        #消耗热量
        txt_burn='消耗热量 '+str(int(data['train']['calories']))+' 千卡'
        
        #教练
        # ins=ins[8:][0]+'教练'
        # print(self.df_ins)
        ins=self.df_ins.loc[self.df_ins['员工编号']==ins[0:8]]['昵称'].values[0]

        #建议
        txt_suggest_title=ins+'给你的饮食建议'
        # txt_suggest='补充足够的碳水化合物：健身训练时能量主要由糖原提供，摄入的碳水化合物可以补充糖原，供给能量，并防止训练造成的肌肉分解'
        exp_knlg_fn=os.path.join(self.exp_knlg_dir,'减脂饮食建议表.xlsx')
        _diet_suggests=get_data.ReadDiet(exp_knlg_fn)
        diet_suggests=_diet_suggests.exp_diet_suggests()
        txt_suggest=random.choice(diet_suggests)

        #slogan
        # txt_slogan='让健身变得有趣'
        txt_slogan= self.txt_slogan.strip()

        # print(nickname,sex,'\n',txt_date,'\n',txt_train,txt_calories,'\n',ins,txt_suggest,slogan)


        ftsz_train=42
        ftsz_suggest=40

        #背景

        def cal_ht():

            size={
                'wid':{
                    'total':720,
                    'small':640,
                    'third':600
                },
                'ht':{
                    'total':1280,
                    'title':300,
                    'title_box':260,
                    'train':300,
                    'burn':200,
                    'suggest':300,
                    'suggest_title':120,
                    'bottom':200,
                    'gap':10
                }
            }
            
            #重写训练内容高度
            ht_train_cal=composing.split_txt_Chn_eng(wid=size['wid']['small']-40,font_size=ftsz_train,txt_input=txt_train,Indent='no')
            # ht_train=800
            ht_train=int(ftsz_train*ht_train_cal[1]*1.9)
            if ht_train>=500:
                ht_train=int(ftsz_train*ht_train_cal[1]*1.55)
            # print(ht_train)
            size['ht']['train']=ht_train

            #重写建议内容高度
            ht_suggest_cal=composing.split_txt_Chn_eng(wid=size['wid']['third']-20,font_size=ftsz_suggest,txt_input=txt_suggest,Indent='yes')
            ht_suggest=int(ftsz_suggest*ht_suggest_cal[1]*2)+size['ht']['suggest_title']
            size['ht']['suggest']=ht_suggest

            total_ht=size['ht']['title']+size['ht']['train']+size['ht']['burn']+size['ht']['suggest']+size['ht']['bottom']+size['ht']['gap']*2*4
            size['ht']['total']=total_ht
 
            # print(size)
            return size 
        
        def color_list():
            color_config=readconfig.exp_json(self.color_config_fn)
            if data['sex']=='女':
                color=color_config['AfterClass']['pink']
            elif data['sex']=='男':
                color=color_config['AfterClass']['blue']
            return color

        def draw_blocks():
            size=cal_ht()
            color=color_list()

            p_title_block=(0,0,size['wid']['total'],size['ht']['title'])
            p_title_box=(p_title_block[0]+(size['wid']['total']-size['wid']['small'])//2,
                        p_title_block[1]+(size['ht']['title']-size['ht']['title_box'])//2,
                        p_title_block[0]+(size['wid']['total']-size['wid']['small'])//2+size['wid']['small'],
                        p_title_block[1]+(size['ht']['title']-size['ht']['title'])//2+size['ht']['title_box'])


            p_train=(p_title_block[0]+(size['wid']['total']-size['wid']['small'])//2,
                        p_title_block[3]+size['ht']['gap']*2,
                        p_title_block[0]+(size['wid']['total']-size['wid']['small'])//2+size['wid']['small'],
                        p_title_block[3]+size['ht']['train'])
            p_train_bar=(p_train[0]+50,p_train[1]+18,p_train[0]+30+10,p_train[3]-15)
            p_train_txt=[p_train_bar[0]+35,p_train_bar[1]+10]
            p_burn=(p_train[0],
                    p_train[3]+size['ht']['gap']*2,
                    p_train[2],
                    p_train[3]+size['ht']['burn'])
            
            p_flame=[p_burn[0]+20,p_burn[1]+30]
                    
            p_suggest=(p_burn[0],
                    p_burn[3]+size['ht']['gap']*2,
                    p_burn[2],
                    p_burn[3]+size['ht']['gap']*2+size['ht']['suggest'])

            p_suggest_small=(p_suggest[0]+(size['wid']['small']-size['wid']['third'])//2,
                    p_suggest[1]+size['ht']['suggest_title'],
                    p_suggest[0]+(size['wid']['small']-size['wid']['third'])//2+size['wid']['third'],
                    p_suggest[1]+size['ht']['suggest_title']+(size['ht']['suggest']-size['ht']['suggest_title'])-20)

            p_suggest_txt=[p_suggest_small[0]+18,p_suggest_small[1]+20]
            
            p_logo=[p_suggest[0]+20,p_suggest[3]+size['ht']['gap']*2+(size['ht']['bottom']-120)//2]
            

            bg=Image.new('RGBA',(size['wid']['total'],size['ht']['total']),color=color['block']['bg'])
            
            draw=ImageDraw.Draw(bg)

            #标题框
            draw.rectangle(p_title_block,fill=color['block']['title'])
            draw.rounded_rectangle(xy=p_title_box,radius=10,fill=None,width=3,outline=color['edge']['title_box'])


            #训练内容框
            draw.rounded_rectangle(xy=p_train,radius=10,fill=color['block']['train'],width=3,outline=None)
            draw.rectangle(xy=p_train_bar,fill=color['block']['train_bar'])

            #燃烧
            draw.rounded_rectangle(xy=p_burn,radius=10,fill=color['block']['burn'],width=3,outline=color['edge']['burn'])

            #教练建议
            draw.rectangle(xy=p_suggest,fill=color['block']['suggest'])
            draw.rounded_rectangle(xy=p_suggest_small,radius=10,fill=color['block']['suggest_small_box'],
                                    width=3,outline=color['edge']['suggest_small_box'])


            #图片
            #火焰图片
            _flame=Image.open(os.path.join(self.public_dir,'flame.png'))
            flame=_flame.resize((_flame.size[0]*120//_flame.size[1],120))
            a_flame=flame.split()[3]
            # bg.paste(a_flame,p_flame)
            bg.paste(flame,p_flame,mask=a_flame)

            #logo
            _logo=Image.open(os.path.join(self.public_dir,'logo及二维码','logo.png'))
            logo=_logo.resize((_logo.size[0]*120//_logo.size[1],120))
            a_logo=logo.split()[3]
            bg.paste(logo,p_logo,mask=a_logo)

            #文字
            font_config_file=os.path.join(os.path.dirname(__file__),'configs','FontList.minghu.config')
            #姓名
            draw.text((p_title_box[0]+50,p_title_box[1]+30),
                        nickname+sex,
                        fill=color['font']['title'],
                        font=composing.fonts('方正韵动粗黑',60,config=font_config_file))
            #日期
            draw.text((p_title_box[0]+50,p_title_box[1]+110),
                        txt_date,
                        fill=color['font']['title'],
                        font=composing.fonts('方正韵动粗黑',40,config=font_config_file))
            #标题栏内其他文字
            draw.text((p_title_box[0]+50,p_title_box[1]+176),
                        txt_title_box,
                        fill=color['font']['title'],
                        font=composing.fonts('方正韵动粗黑',40,config=font_config_file))

            #训练内容
            print()
            composing.put_txt_img(draw=draw,
                                    tt=txt_train,
                                    total_dis=int((p_train[2]-p_train_bar[0])*0.8),
                                    xy=p_train_txt,
                                    dis_line=int(ftsz_train*0.5),
                                    fill=color['font']['train'],
                                    font_name='汉仪糯米团',
                                    font_size=ftsz_train,
                                    addSPC='no',
                                    font_config_file=font_config_file)

            #燃烧热量
            draw.text((p_burn[0]+130,p_burn[1]+52),
                        txt_burn,
                        fill=color['font']['burn'],
                        font=composing.fonts('汉仪糯米团',54,config=font_config_file))

            #教练建议
            draw.text((p_suggest[0]+98,p_suggest[1]+36),
                        txt_suggest_title,
                        fill=color['font']['suggest_title'],
                        font=composing.fonts('汉仪糯米团',44,config=font_config_file))

            composing.put_txt_img(draw=draw,
                                    tt=txt_suggest,
                                    total_dis=int((p_suggest_small[2]-p_suggest_small[0])*0.9),
                                    xy=p_suggest_txt,
                                    dis_line=int(ftsz_suggest*0.5),
                                    fill=color['font']['suggest'],
                                    font_name='汉仪字酷堂义山楷w',
                                    font_size=ftsz_train,
                                    addSPC='yes',
                                    font_config_file=font_config_file)

            #slogan
            draw.text((p_logo[0]+255,p_logo[1]+28),txt_slogan,
                        fill=color['font']['slogan'],font=composing.fonts('华康海报体W12(p)',52,config=font_config_file))

            # bg.show()
            bg=bg.convert('RGB')
            save_name=date_input+'_'+cus+'.jpg'
            save_dir=os.path.join(self.save_dir_feedback,cus)
            if not os.path.exists(save_dir):
                os.makedirs(save_dir)
            bg.save(os.path.join(save_dir,save_name),quality=90,subsampling=0)
            
            if open_dir=='yes':
                os.startfile(save_dir)


            print('完成\n')


        draw_blocks()


    def draw_new(self,cus='MH024刘婵桢',ins='MHINS002韦越棋',date_input='20210324',open_dir='yes'):
        #公共文字
        # with open(os.path.join(self.material_dir,'txt_public.txt'),'r',encoding='utf-8') as txt_pub:
        #     txt_public=txt_pub.readlines()

        #文字内容
        data=self.export(cus=cus,ins=ins,date_input=date_input)
        # print(data)
        #日期
        txt_date=date_input[:4]+'年'+date_input[4:6]+'月'+date_input[6:]+'日'
        #姓名
        nickname=data['nickname']
        #性别
        sex=data['sex']
        if sex=='女':
            sex='女士'
        elif sex=='男':
            sex='先生'
        else:
            sex=''
        
        #标题框文字
        # txt_title_box='看看今天你的汗水洒在哪里？'
        txt_title_box=self.txt_mini_title
        

        #抗阻内容
        txt_train_muscle=''
        for mscl_item in data['train']['muscle_item']:
            if mscl_item[2]>0:
                txt_train_muscle=txt_train_muscle+mscl_item[0]+'  '+str(int(mscl_item[2]))+'个'+'\n'
            else:
                txt_train_muscle=txt_train_muscle+mscl_item[0]+'  '+str(int(mscl_item[3]))+'米'+'\n'
        txt_train_muscle.strip()

        #有氧内容
        txt_train_oxy=''
        for oxy_item in  data['train']['oxy_infos']:
            if oxy_item[1]%60==0:
                txt_train_oxy=txt_train_oxy+oxy_item[0]+'  '+str(int(oxy_item[1]//60))+'分'+'\n'
            else:
                # txt_train_oxy=txt_train_oxy+oxy_item[0]+'  '+str(int(oxy_item[1]//60))+'分'+'\n'
                txt_train_oxy=txt_train_oxy+oxy_item[0]+'  '+str(int(oxy_item[1]//60))+'分'+str(int(oxy_item[1]%60))+'秒\n'
        txt_train_oxy.strip()


        txt_train=txt_train_muscle+txt_train_oxy

        #消耗热量
        txt_burn='消耗热量 '+str(int(data['train']['calories']))+' 千卡'
        
        #教练
        # ins=ins[8:][0]+'教练'
        # print(self.df_ins)
        ins=self.df_ins.loc[self.df_ins['员工编号']==ins[0:8]]['昵称'].values[0]

        #建议
        txt_suggest_title=ins+'给你的饮食建议'
        # txt_suggest='补充足够的碳水化合物：健身训练时能量主要由糖原提供，摄入的碳水化合物可以补充糖原，供给能量，并防止训练造成的肌肉分解'
        exp_knlg_fn=os.path.join(self.exp_knlg_dir,'减脂饮食建议表.xlsx')
        _diet_suggests=get_data.ReadDiet(exp_knlg_fn)
        diet_suggests=_diet_suggests.exp_diet_suggests()
        txt_suggest=random.choice(diet_suggests)

        #slogan
        # txt_slogan='让健身变得有趣'
        txt_slogan= self.txt_slogan.strip()

        # print(nickname,sex,'\n',txt_date,'\n',txt_train,txt_calories,'\n',ins,txt_suggest,slogan)


        ftsz_train=40
        ftsz_suggest=40

        #背景

        def cal_ht():

            size={
                'wid':{
                    'total':720,
                    'small':640,
                    'third':600
                },
                'ht':{
                    'total':1280,
                    'title':300,
                    'title_box':230,
                    'train':350,
                    'burn':200,
                    'suggest':300,
                    'suggest_title':120,
                    'bottom':200,
                    'gap':10
                }
            }
            
            #重写训练内容高度
            ht_train_cal=composing.split_txt_Chn_eng(wid=size['wid']['small']-40,font_size=ftsz_train,txt_input=txt_train,Indent='no')
            # ht_train=800
            ht_train=int(ftsz_train*ht_train_cal[1]*2.2)
            if ht_train>=500:
                ht_train=int(ftsz_train*ht_train_cal[1]*1.8)
            # print(ht_train)
            size['ht']['train']=ht_train

            #重写建议内容高度
            ht_suggest_cal=composing.split_txt_Chn_eng(wid=size['wid']['third']-20,font_size=ftsz_suggest,txt_input=txt_suggest,Indent='yes')
            ht_suggest=int(ftsz_suggest*ht_suggest_cal[1]*2)+size['ht']['suggest_title']
            size['ht']['suggest']=ht_suggest

            total_ht=size['ht']['title']+size['ht']['train']+size['ht']['burn']+size['ht']['suggest']+size['ht']['bottom']+size['ht']['gap']*2*4
            size['ht']['total']=total_ht
 
            # print(size)
            return size 
        
        def color_list():
            color_config=readconfig.exp_json(self.color_config_fn)
            if data['sex']=='女':
                color=color_config['AfterClass']['grey']
            elif data['sex']=='男':
                color=color_config['AfterClass']['blue']
            return color

        def draw_blocks():
            size=cal_ht()
            color=color_list()

            p_title_block=(0,0,size['wid']['total'],size['ht']['title'])
            p_title_box=(p_title_block[0]+(size['wid']['total']-size['wid']['small'])//2,
                        p_title_block[1]+(size['ht']['title']-size['ht']['title_box'])//2,
                        p_title_block[0]+(size['wid']['total']-size['wid']['small'])//2+size['wid']['small'],
                        p_title_block[1]+(size['ht']['title']-size['ht']['title'])//2+size['ht']['title_box'])


            p_train=(p_title_block[0]+(size['wid']['total']-size['wid']['small'])//2,
                        p_title_block[3]+size['ht']['gap']*2,
                        p_title_block[0]+(size['wid']['total']-size['wid']['small'])//2+size['wid']['small'],
                        p_title_block[3]+size['ht']['train'])
            p_train_bar=(p_train[0]+110,p_train[1]+26,p_train[0]+110+8,p_train[3]-18)
            p_train_txt=[p_train_bar[0]+50,p_train_bar[1]+4]
            p_burn=(p_train[0],
                    p_train[3]+size['ht']['gap']*2,
                    p_train[2],
                    p_train[3]+size['ht']['burn'])
            
            p_flame=[p_burn[0]+30,p_burn[1]+50]
                    
            p_suggest=(p_burn[0],
                    p_burn[3]+size['ht']['gap']*2,
                    p_burn[2],
                    p_burn[3]+size['ht']['gap']*2+size['ht']['suggest'])

            p_suggest_small=(p_suggest[0]+(size['wid']['small']-size['wid']['third'])//2,
                    p_suggest[1]+size['ht']['suggest_title'],
                    p_suggest[0]+(size['wid']['small']-size['wid']['third'])//2+size['wid']['third'],
                    p_suggest[1]+size['ht']['suggest_title']+(size['ht']['suggest']-size['ht']['suggest_title'])-20)

            p_suggest_txt=[p_suggest_small[0]+18,p_suggest_small[1]+20]
            
            p_logo=[p_suggest[0]+20,p_suggest[3]+size['ht']['gap']*2+(size['ht']['bottom']-120)//2]
            

            bg=Image.new('RGBA',(size['wid']['total'],size['ht']['total']),color=color['block']['bg'])
            
            draw=ImageDraw.Draw(bg)

            #标题框
            # draw.rectangle(p_title_block,fill=color['block']['title'])
            draw.rounded_rectangle(xy=p_title_box,radius=10,fill=color['edge']['title_box'],width=3,outline=color['edge']['title_box'])


            #训练内容框
            draw.rounded_rectangle(xy=p_train,radius=10,fill=color['block']['train'],width=3,outline=None)
            draw.rectangle(xy=p_train_bar,fill=color['block']['train_bar'])

            #燃烧
            draw.rounded_rectangle(xy=p_burn,radius=10,fill=color['block']['burn'],width=3,outline=color['edge']['burn'])

            #教练建议
            draw.rectangle(xy=p_suggest,fill=color['block']['suggest'])
            draw.rounded_rectangle(xy=p_suggest_small,radius=10,fill=color['block']['suggest_small_box'],
                                    width=3,outline=color['edge']['suggest_small_box'])


            #图片
            #火焰图片
            # _flame=Image.open(os.path.join(self.public_dir,'flame.png'))
            # flame=_flame.resize((_flame.size[0]*120//_flame.size[1],120))
            # a_flame=flame.split()[3]
            # # bg.paste(a_flame,p_flame)
            # bg.paste(flame,p_flame,mask=a_flame)
            flame_ico=Image.open(os.path.join(self.public_dir,'UI图标','calory02.png'))
            flame_ico=self.icon(flame_ico,ico_size=(60,60))
            # flame_ico=Image.open(os.path.join(self.public_dir,'flame.png'))
            # flame_ico=self.icon(flame_ico,ico_size=(flame_ico.size[0]*60//flame_ico.size[1],60))
            bg.paste(flame_ico[0],p_flame,mask=flame_ico[1])

            #头像    
            if sex=='女士':
                head_ico=Image.open(os.path.join(self.public_dir,'UI图标','head_female.png'))
            else:
                head_ico=Image.open(os.path.join(self.public_dir,'UI图标','head_male.png'))
            head_ico=self.icon(head_ico,ico_size=(50,50))            
            bg.paste(head_ico[0],(p_title_box[0]+36,p_title_box[1]+30),mask=head_ico[1])

            #日记
            diary_ico=Image.open(os.path.join(self.public_dir,'UI图标','calendar.png'))
            diary_ico=self.icon(diary_ico,ico_size=(50,50))
            bg.paste(diary_ico[0],(p_title_box[0]+36,p_title_box[1]+106),mask=diary_ico[1])


            

            #logo
            _logo=Image.open(os.path.join(self.public_dir,'logo及二维码','logo.png'))
            logo=_logo.resize((_logo.size[0]*160//_logo.size[1],160))
            a_logo=logo.split()[3]
            bg.paste(logo,p_logo,mask=a_logo)

            #文字
            font_config_file=os.path.join(os.path.dirname(__file__),'configs','FontList.minghu.config')
            #姓名
            draw.text((p_title_box[0]+116,p_title_box[1]+30),
                        nickname,
                        fill=color['font']['title'],
                        font=composing.fonts('方正韵动粗黑',50,config=font_config_file))
            draw.text((p_title_box[0]+116+180,p_title_box[1]+45),
                        sex,
                        fill=color['font']['title'],
                        font=composing.fonts('思源黑体',30,config=font_config_file))
            #日期
            draw.text((p_title_box[0]+116,p_title_box[1]+110),
                        txt_date+'   训练日记',
                        fill=color['font']['title'],
                        font=composing.fonts('思源黑体',40,config=font_config_file))
            #标题栏内其他文字
            draw.text((p_title_box[0]+116,p_title_box[1]+206),
                        txt_title_box,
                        fill=color['font']['title'],
                        font=composing.fonts('思源黑体',40,config=font_config_file))

            #训练内容
            print()
            composing.put_txt_img(draw=draw,
                                    tt=txt_train,
                                    total_dis=int((p_train[2]-p_train_bar[0])*0.8),
                                    xy=p_train_txt,
                                    dis_line=int(ftsz_train*0.6),
                                    fill=color['font']['train'],
                                    font_name='字由文艺黑体',
                                    font_size=ftsz_train,
                                    addSPC='no',
                                    font_config_file=font_config_file)

            #燃烧热量
            draw.text((p_burn[0]+130,p_burn[1]+52),
                        txt_burn,
                        fill=color['font']['burn'],
                        font=composing.fonts('汉仪糯米团',54,config=font_config_file))

            #教练建议
            _ico_dot=Image.open(os.path.join(self.public_dir,'UI图标','dot.png'))
            ico_dot=self.icon(_ico_dot,ico_size=(40,40))
            bg.paste(ico_dot[0],(p_suggest[0]+10,p_suggest[1]+56),mask=ico_dot[1])
            draw.line((p_suggest[0]+70,p_suggest[1]+100,p_suggest[0]+40+540,p_suggest[1]+100),fill='#787878')
            draw.text((p_suggest[0]+68,p_suggest[1]+38),
                        txt_suggest_title,
                        fill=color['font']['suggest_title'],
                        font=composing.fonts('思源黑体',44,config=font_config_file))

            composing.put_txt_img(draw=draw,
                                    tt=txt_suggest,
                                    total_dis=int((p_suggest_small[2]-p_suggest_small[0])*0.9),
                                    xy=p_suggest_txt,
                                    dis_line=int(ftsz_suggest*0.5),
                                    fill=color['font']['suggest'],
                                    font_name='汉仪字酷堂义山楷w',
                                    font_size=ftsz_train,
                                    addSPC='yes',
                                    font_config_file=font_config_file)

            #slogan
            draw.text((p_logo[0]+185,p_logo[1]+48),txt_slogan,
                        fill=color['font']['slogan'],font=composing.fonts('华康海报体W12(p)',62,config=font_config_file))

            # bg.show()
            bg=bg.convert('RGB')
            # bg.show()
            save_name=date_input+'_'+cus+'.jpg'
            save_dir=os.path.join(self.save_dir_feedback,cus)
            if not os.path.exists(save_dir):
                os.makedirs(save_dir)
            bg.save(os.path.join(save_dir,save_name),quality=90,subsampling=0)
            
            if open_dir=='yes':
                os.startfile(save_dir)


            print('完成\n')


        draw_blocks()

    def group_afterclass(self,ins='MHINS002韦越棋',date_input='20210324',open_dir='no'):
        grp_file=os.path.join(self.cus_file_dir,'00-团课分班录入表.xlsx')
        df_grp_pre=pd.read_excel(grp_file,sheet_name='分组',skiprows=1)
        df_grp=df_grp_pre.iloc[:,4:]
        df_grp_names=df_grp.columns.tolist()
        #需录入数据的名单
        df_real_list=df_grp_pre['Unnamed: 0'].dropna()           

        if df_real_list.empty:
            exit('未录入数据')
        else:
            # cus_list=df_real_list.apply(lambda x:x+'.xlsx').tolist()
            cus_list=df_real_list.tolist()

        for cus_name in cus_list:
            fn=os.path.join(self.cus_file_dir,cus_name)    
            print('正在生成 {} 的课后反馈……'.format(cus_name),end='')
            self.draw(cus=cus_name,ins=ins,date_input=date_input,open_dir=open_dir)
  
class Vividict(dict):
    def __missing__(self, key):
        value = self[key] = type(self)()
        return value

class PeroidSummary:
    def __init__(self,place='minghu',adj_bfr='yes',adj_src='prg',gui=''):
        self.dir=os.path.dirname(os.path.abspath(__file__))
        config=readconfig.exp_json(os.path.join(self.dir,'configs','main_'+place+'.config'))
        self.cus_file_dir=config['会员档案文件夹']
        self.material_dir=config['素材文件夹']
        self.ins_dir=config['教练文件夹']
        self.slogan_dir=config['文案文件夹']
        self.save_dir=config['输出文件夹']
        self.public_dir=config['公共素材文件夹']
        self.font_dir=config['字体文件夹']
        self.pro_dir=config['专业资料文件夹']
        self.adj_bfr=adj_bfr
        self.adj_src=adj_src
        self.gui=gui
        self.place=place
        # print(os.path.join(self.dir,'configs','main_'+place+'.config'),self.ins_dir,self.cus_file_dir)
        self.df_ins=pd.read_excel(os.path.join(self.ins_dir,'教练信息.xlsx'),sheet_name='教练信息')
        self.color_config_fn=os.path.join(os.path.dirname(__file__),'configs','colors.config')
        with open(os.path.join(self.material_dir,'txt_public.txt'),'r',encoding='utf-8') as txt_pub:
            self.txt_public=txt_pub.readlines()
        self.cus_instance_name=self.txt_public[5].strip()[0:2]
        self.prefix=self.cus_instance_name[0:2]
        self.gym_name=self.txt_public[4]
        self.gym_addr=self.txt_public[3]
        if '%' in self.gym_addr:
            self.gym_addr=''
        self.txt_ins_word=self.txt_public[2]
        self.txt_mini_title=self.txt_public[1]
        self.txt_slogan=self.txt_public[0]

    def fonts(self,font_name,font_size):
        fontList=readconfig.exp_json(os.path.join(self.dir,'configs','FontList.minghu.config'))
        # print(fontList)
        return ImageFont.truetype(fontList[font_name],font_size)

    def color_bg(self,theme='lightgrey'):
        color=readconfig.exp_json(os.path.join(os.path.dirname(__file__),'configs','colors.config'))
        return color['PeriodSummary'][theme]

    def cal_data(self,cus_name_input='MH003吕雅颖',start_date='20210729',end_date='20220201',bmi_bg='#ffffff',bfr_bg="#ffffff",radar_bg='#ffffff',msr_chart_bg='#ffffff'):
        df_basic=pd.read_excel(os.path.join(self.cus_file_dir,cus_name_input+'.xlsx'),sheet_name='基本情况')
        cus_name=df_basic['姓名'].tolist()[0]
        cus_nickname=df_basic['昵称'].tolist()[0]
        cus_sex=df_basic['性别'].tolist()[0]
        #称呼
        if cus_sex=='女':
            txt_cus_sex='女士'
        else:
            txt_cus_sex='先生'
        cus_birthday=df_basic['出生年月'].tolist()[0]
        #起止日期
        txt_period=start_date[:4]+'年'+start_date[4:6]+'月'+start_date[6:]+'日 — '+end_date[:4]+'年'+end_date[4:6]+'月'+end_date[6:]+'日'

        #标准化生日
        if len(str(cus_birthday))==4:
            cus_birthday=int(str(cus_birthday)+'0101')
        elif len(str(cus_birthday))==6:
            cus_birthday=int(str(cus_birthday)+'01')

        e_date=datetime(int(end_date[0:4]),int(end_date[4:6]),int(end_date[6:8]))
        s_date=datetime(int(start_date[0:4]),int(start_date[4:6]),int(start_date[6:8]))
        interval=e_date-s_date
        prd=interval.days

        dif_y,dif_m,dif_d=days_cal.Dates().dif_y_m_d(s=start_date,e=end_date)
        if dif_y!=0:
            txt_dif_y=str(dif_y)+'年'            
        else:
            txt_dif_y=''

        if dif_m!=0:
            txt_dif_m=str(dif_m)+'个月'
        else:
            txt_dif_m=''

        if dif_d!=0:
            txt_dif_d=str(dif_d)+'天'
        else:
            txt_dif_d=''

        _txt_dif=[txt_dif_y,txt_dif_m,txt_dif_d]
        _txt_dif=list(filter(None,_txt_dif))
        if len(_txt_dif)>1:
            _txt_dif[-1]='零'+_txt_dif[-1]
        txt_prd_0=''.join(_txt_dif)


        df_train=pd.read_excel(os.path.join(self.cus_file_dir,cus_name_input+'.xlsx'),sheet_name='训练情况',skiprows=1)
        df_train.columns=['时间','形式','目标肌群','有氧项目','有氧时长','抗阻内容','重量','距离','次数','消耗热量','教练姓名','教练评语']
        df_train_interval=df_train[(df_train['时间']>=s_date) & (df_train['时间']<=e_date) ]
        df_ym=pd.DataFrame()
        df_ym['year']=df_train_interval['时间'].dt.year
        df_ym['month']=df_train_interval['时间'].dt.month
        df_ym['date']=df_train_interval['时间']
        df_ym_count=df_ym.drop_duplicates('month')
        df_ym_cal=df_ym_count.groupby('year').count().reset_index()
        months=df_ym_cal['month'].sum()

        if months>12:
            txt_prd=str(months//12)+'年零'+str(months//12)+'个月'
        else:
            txt_prd=str(months)+'个月'
        
        
        #训练次数（唯一的日期计数）
        train_counts=len(list(df_train_interval['时间'].unique()))

        #平均训练频率        
        # if train_counts//months>=1:
        #     avr_train_counts=str(train_counts//months)+'次'
        # else:
        #     avr_train_counts='每月不到1次'
        try:
            if train_counts//(prd//30)>=1:
                txt_avr_train_counts='你平均每个月运动 '+str(train_counts//(prd//30))+'次。'
            else:
                txt_avr_train_counts='你每个月运动不到1次。'
        except:
            txt_avr_train_counts='每个月运动'+str(train_counts)+'次。'

            # '在过去的 '+contents['train_time']+' 里，你平均每个月运动 '+str(contents['train_frqcy'])+'。'

        #最高训练频率及对应月份
        df_ym_trainmax=df_ym.drop_duplicates('date')
        df_trainmax=df_ym_trainmax.groupby(['year','month']).count().reset_index()
        trainmax=df_trainmax[df_trainmax['date']==df_trainmax['date'].max()]
        if trainmax.shape[0]>1:
            t_month=''
            for index,row in trainmax.iterrows():
                t_month=t_month+str(row['year'])+'年'+str(row['month'])+'月、'
            _txt_trainmax=t_month[:-1]+'，这'+str(trainmax.shape[0])+'个月里每个月都运动了'+str(trainmax['date'].max())+'次。'
        else:
            _txt_trainmax='表现最优秀是在 '+str(trainmax['year'].tolist()[0])+'年'+str(trainmax['month'].tolist()[0])+'月，运动了'+str(trainmax['date'].max())+'次。'
        
        #最高训练文字太长则分段
        if len(_txt_trainmax)>33:
            txt_trainmax=_txt_trainmax[:33]+'\n\n'+_txt_trainmax[33:]
        else:
            txt_trainmax=_txt_trainmax


        #最后一次体测日期
        
        df_measure=pd.read_excel(os.path.join(self.cus_file_dir,cus_name_input+'.xlsx'),sheet_name='身体数据',skiprows=0)
        txt_latest_msr_date=str(df_measure['时间'].max())[0:4]+'年'+str(df_measure['时间'].max())[5:7]+'月'+str(df_measure['时间'].max())[8:10]+'日'

        #体重
        wt=df_measure[df_measure['时间']==df_measure['时间'].max()]['体重'].tolist()[0]
        txt_wt=str(wt)+' Kg'
        ht=df_measure[df_measure['时间']==df_measure['时间'].max()]['身高'].tolist()[0]

        if np.isnan(ht) or np.isnan(wt) :
            print('身体数据有未填写项，请核实。')
            # exit(0)
            return
            
        #BMI        
        txt_bmi=str(round(wt/((ht/100)*(ht/100)),2))
        bmi_chart=draw_pic.Scale(scale_name='BMI',stage=[10,18.5,24,28,40],stage_name=['','','超重','肥胖',''],colors=('#9ED6D6','#CEE9E9','#CEE9E9','#9ED6D6','#9ED6D6'))
        pic_bmi=bmi_chart.draw(val=round(wt/((ht/100)*(ht/100)),2),color_val='#84B6B9',scale_adj=200,color_bg=bmi_bg,back_transparent_color='',arrow_fn=os.path.join(self.public_dir,'UI图标','倒三角_blue.png'))

        
        cals=get_data.cals()
        age=days_cal.calculate_age(str(cus_birthday))

        #BMR
        txt_bmr=str(round(cals.bmr(sex=cus_sex,ht=ht,wt=wt,age=age),2))+' 千卡'


        #BFR
        cus_waist=df_measure[df_measure['时间']==df_measure['时间'].max()]['腰围'].tolist()[0]
        val_bfr=round(cals.bfr(age=age,sex=cus_sex,ht=ht,wt=wt,waist=cus_waist,adj_bfr=self.adj_bfr,adj_src=self.adj_src,gui=self.gui,formula=1)*100,2)
        txt_bfr=str(val_bfr)+' %'
        if cus_sex=='女':
            bfr_stage=[10,25,28,32,40]
            stage_name=['','','丰满','肥胖','']
            scale_stage=300
        else:
            bfr_stage=[0,15,18,25,30]
            stage_name=['','腹肌\n清晰','腹肌\n隐约','肥胖','']
            scale_stage=100
        bfr_chart=draw_pic.Scale(scale_name='BFR',stage=bfr_stage,stage_name=stage_name,colors=('#FBF2DA','#FBF2DA','#FBF2DA','#FBF2DA','#EDD9A5','#EDD9A5'))
        pic_bfr=bfr_chart.draw(val=val_bfr,color_val='#CEC09C',scale_adj=scale_stage,color_bg=bfr_bg,back_transparent_color='',arrow_fn=os.path.join(self.public_dir,'UI图标','倒三角_yellow.png'))



        #训练数据
        train_data=get_data.ReadAndExportDataNew(adj_bfr='no',adj_src='prg',gui='').exp_cus_prd(self.cus_file_dir,cus=cus_name_input,start_time=start_date,end_time=end_date)


        #有氧训练时长
        oxy_time=train_data['train']['oxy_time']
        if oxy_time>86400:
            txt_oxy_time=str(int(oxy_time//86400))+'天'+str(int(oxy_time%86400//3600))+'小时'
        elif oxy_time>3600:
            txt_oxy_time=str(int(oxy_time//3600))+'小时'+str(int(oxy_time%3600//60))+'分'
        else:
            txt_oxy_time=str(int(oxy_time//60))+'分'

        #抗阻训练总重量
        txt_muscle_wt=str(int(train_data['train']['muscle_total_wt']))+' Kg'

        #各部位训练次数
        each_part=train_data['train']['muscle']
        txt_each_part=''
        for itm in each_part:
            txt_each_part=txt_each_part+'     -  '+itm+':  '+str(each_part[itm])+' 次'+'\n\n'
    
        #运动消耗
        txt_calories=str(int(train_data['train']['calories']))+' Kcal'

        #体适能指标
        # print(train_data['body']['ht_lung'])
        physical_fitness_data={'心肺':train_data['body']['ht_lung'],
                                '平衡':train_data['body']['balance'],
                                '力量':train_data['body']['power'],
                                '柔韧性':train_data['body']['flexibility'],
                                '核心':train_data['body']['core']}
        
        radar=draw_pic.DrawRadar()
        pic_radar=radar.draw(physical_fitness_data,bgcolor=radar_bg)
        # pic_radar.show()


        #围度变化曲线
    
        body_measure_data=draw_pic.PeriodChart(font_fn=os.path.join(self.font_dir,'msyh.ttc'))
        body_measure_chart=body_measure_data.to_pic(cus_dir=self.cus_file_dir,cus_fn=cus_name_input+'.xlsx',start_time=start_date,end_time=end_date,d_font='',title='',bgcolor=msr_chart_bg,items=['waist','hip','chest'])
        # body_measure_chart.show()
                                

        # print(latest_msr_date,txt_wt,txt_bmi,txt_bmr,txt_bfr)
        # print(txt_calories)
        contents={'nickname':cus_nickname,'sex':cus_sex,'s-e_date':txt_period,
                    'train_time':txt_prd_0,'train_frqcy':txt_avr_train_counts,
                    'train_max_frqcy':txt_trainmax,'latest_msr':txt_latest_msr_date,
                    'wt':txt_wt,'bmi':txt_bmi,'bmr':txt_bmr,'bfr':txt_bfr,
                    'oxy_time':txt_oxy_time,'muscle_wt':txt_muscle_wt,'each_part':txt_each_part,
                    'calories':txt_calories,'pic_radar':pic_radar,'pic_msr_chart':body_measure_chart,
                    'pic_bmi':pic_bmi,'pic_bfr':pic_bfr}

        return contents

    def read_diet(self):
        df=pd.read_excel(os.path.join(self.pro_dir,'减脂饮食建议表.xlsx'),sheet_name='饮食建议')
        # df_sgst=df['饮食建议'].tolist()
        df.dropna(axis=0,how='any',inplace=True)
        df_sgst=df['饮食建议'].sample(1).tolist()[0]
        return df_sgst

    def diet_txts(self,wid=680,font_size=28):
        txt_input=self.read_diet()
        para_txt=composing.split_txt_Chn_eng(wid=wid,font_size=font_size,txt_input=txt_input,Indent='yes')
        # print('main文字段数',para_txt[1])
        return para_txt

    def block_ht(self,contents_input,diet_para_num,diary_font_size=28,diet_font_size=28):
        title=120
        basic_info=330
        basic_body=1540
        t_diary=contents_input['each_part'].split('\n')
        t_diary.append(contents_input['muscle_wt'].split('\n'))
        t_diary.append(contents_input['oxy_time'].split('\n'))
        #根据运动日记长短调整色块高度
        if len(t_diary)*diary_font_size>=300:
            diary=len(t_diary)*diary_font_size*2
        else:
            diary=len(t_diary)*diary_font_size*2-int((4*len(t_diary)*diary_font_size*2)/3)+440
        # print(-int((4*len(t_diary)*diary_font_size*2)/3)+440,len(t_diary)*diary_font_size*2)
        msr_change=900
        diet=math.ceil(diet_para_num*2.4*diet_font_size)+90
        bottom=150
        content=[title,basic_info,basic_body,diary,msr_change,diet,bottom]        
        gap=20
        total_ht=sum(content)+gap*(len(content)-1)

        return {'b_title':title,'b_info':basic_info,'b_body':basic_body,
                'b_diary':diary,'b_msr':msr_change,'b_diet':diet,'b_bottom':bottom,
                'gap':gap,'total_ht':total_ht}

    def icon(self,ico,ico_size):
        ico=ico.resize(ico_size)
        return (ico,ico.split()[3])

    def exp_chart(self,cus_name_input='MH003吕雅颖',ins='MHINS001陆伟杰',start_date='20210729',end_date='20220201',
                                theme='lightgrey',ico_size=(40,40),diary_font_size=26,diet_font_size=26,diet_boxwid=580,logo_ht=52):
        colors=self.color_bg(theme=theme)
        contents=self.cal_data(cus_name_input=cus_name_input,start_date=start_date,end_date=end_date,bmi_bg=colors['bmi_bg'],bfr_bg=colors['bfr_bg'],radar_bg=colors['radar_bg'],msr_chart_bg=colors['msr_chart_bg'])
        diet_para_nums=self.diet_txts(wid=diet_boxwid,font_size=diet_font_size)[1]
        
        # print(self.diet_txts(wid=680,font_size=diet_font_size)[0],diet_para_nums,math.ceil(diet_para_nums*2.4*diet_font_size)+90)
        
        block_ht=self.block_ht(contents_input=contents,diet_para_num=diet_para_nums,diary_font_size=diary_font_size,diet_font_size=diet_font_size)
        bg=Image.new('RGBA',(720,block_ht['total_ht']),color=colors['bg'])
        gap=block_ht['gap']

        #坐标计算
        w_block=bg.size[0]

        y_title=0
        y_info=y_title+block_ht['b_title']+gap
        y_body=y_info+block_ht['b_info']+gap
        y_diary=y_body+block_ht['b_body']+gap
        y_msr=y_diary+block_ht['b_diary']+gap
        y_diet=y_msr+block_ht['b_msr']+gap
        y_bottom=y_diet+block_ht['b_diet']+gap

        # print(y_diet,y_bottom,y_msr,block_ht['b_diet'],block_ht['b_bottom'],block_ht['total_ht'])

        draw=ImageDraw.Draw(bg)
        #标题----------------------------------------------------------------------------
        bg_title=Image.new('RGBA',(720,block_ht['b_title']),color=colors['title'])
        bg.paste(bg_title,(0,y_title))        

        logo=Image.open(os.path.join(self.material_dir,'logo及二维码','logo.png'))
        logo=logo.resize((logo.size[0]*logo_ht//logo.size[1],logo_ht))
        m_logo=logo.split()[3]
        bg.paste(logo,(50,30),mask=m_logo)
        draw.text((152,40),self.gym_name.strip()+'会员运动记录',fill='#969696',font=self.fonts('字由文艺黑体',40))

        #基本信息---------------------------------------------------------------
        bg_info=Image.new('RGBA',(720,block_ht['b_info']),color=colors['basic_info'])
        bg.paste(bg_info,(0,y_info))
        #姓名
        if contents['sex']=='女':
            ico_head=Image.open(os.path.join(self.material_dir,'UI图标','head_female.png'))
            txt_sex='女士'
        else:
            ico_head=Image.open(os.path.join(self.material_dir,'UI图标','head_male.png'))
            txt_sex='先生'
        ico_head=self.icon(ico_head,ico_size)
        bg.paste(ico_head[0],(50,180),mask=ico_head[1])
        draw.text((120,178),contents['nickname']+' '+txt_sex,fill='#787878',font=self.fonts('思源黑体',44))

        #训练时间段
        _ico_prd=Image.open(os.path.join(self.material_dir,'UI图标','calendar.png'))
        ico_prd=self.icon(_ico_prd,ico_size)
        bg.paste(ico_prd[0],(50,260),mask=ico_prd[1])
        draw.text((120,270),contents['s-e_date'],fill='#787878',font=self.fonts('思源黑体',22))

        #训练频率
        _ico_frqcy=Image.open(os.path.join(self.material_dir,'UI图标','frequency.png'))
        ico_frqcy=self.icon(_ico_frqcy,ico_size)
        bg.paste(ico_frqcy[0],(50,320),mask=ico_frqcy[1])
        txt_avr_frqcy='在过去的 '+contents['train_time']+' 里，'+str(contents['train_frqcy'])
        draw.text((120,330),txt_avr_frqcy,fill='#787878',font=self.fonts('思源黑体',22))

        #最大训练频率
        _ico_frqcy_max=Image.open(os.path.join(self.material_dir,'UI图标','frequency_max.png'))
        ico_frqcy_max=self.icon(_ico_frqcy_max,ico_size)
        bg.paste(ico_frqcy_max[0],(50,380),mask=ico_frqcy_max[1])
        draw.text((120,390),contents['train_max_frqcy'],fill='#787878',font=self.fonts('思源黑体',22))

        #最大训练频率
        _ico_frqcy_max=Image.open(os.path.join(self.material_dir,'UI图标','frequency_max.png'))
        ico_frqcy_max=self.icon(_ico_frqcy_max,ico_size)
        bg.paste(ico_frqcy_max[0],(50,380),mask=ico_frqcy_max[1])
        draw.text((120,390),contents['train_max_frqcy'],fill='#787878',font=self.fonts('思源黑体',22))

        #基本体格------------------------------------------------------------------------------------------
        bg_body=Image.new('RGBA',(720,block_ht['b_body']),color=colors['basic_body'])
        bg.paste(bg_body,(0,y_body))

        #基本体格标题
        _ico_dot=Image.open(os.path.join(self.material_dir,'UI图标','dot.png'))
        ico_dot=self.icon(_ico_dot,ico_size)
        bg.paste(ico_dot[0],(50,540),mask=ico_dot[1])
        draw.text((100,540),'基本体格',fill='#787878',font=self.fonts('思源黑体',36))
        draw.line((50,590,680,590),fill='#787878')

        #最后测量日期
        _ico_msr=Image.open(os.path.join(self.material_dir,'UI图标','clock.png'))
        ico_msr=self.icon(_ico_msr,ico_size)
        bg.paste(ico_msr[0],(50,620),mask=ico_msr[1])
        draw.text((120,630),'最近测量日期：'+contents['latest_msr'],fill='#787878',font=self.fonts('思源黑体',26))

        #基础代谢率
        _ico_bfr=Image.open(os.path.join(self.material_dir,'UI图标','calory.png'))
        ico_bfr=self.icon(_ico_bfr,ico_size)
        bg.paste(ico_bfr[0],(50,680),mask=ico_bfr[1])
        draw.text((120,690),'基础代谢率：'+contents['bmr'],fill='#787878',font=self.fonts('思源黑体',26))

        #体重/BMI
        _ico_wt=Image.open(os.path.join(self.material_dir,'UI图标','weight.png'))
        ico_wt=self.icon(_ico_wt,ico_size)
        bg.paste(ico_wt[0],(50,760),mask=ico_wt[1])
        draw.text((120,765),'体重：'+contents['wt'],fill='#787878',font=self.fonts('思源黑体',26))
        draw.text((400,765),'BMI：'+contents['bmi'],fill='#787878',font=self.fonts('思源黑体',26))
        #BMI图
        pic_bmi=contents['pic_bmi']
        pic_bmi=pic_bmi.resize((600,600*pic_bmi.size[1]//pic_bmi.size[0]))
        pic_bmi=pic_bmi.crop((0,int(600*pic_bmi.size[1]//pic_bmi.size[0]//3),pic_bmi.size[0],pic_bmi.size[1]))
        bg.paste(pic_bmi,(60,820))
        #体脂率
        _ico_bfr=Image.open(os.path.join(self.material_dir,'UI图标','bfr.png'))
        ico_bfr=self.icon(_ico_bfr,ico_size)
        bg.paste(ico_bfr[0],(50,1070),mask=ico_bfr[1])
        draw.text((120,1080),'体脂率：'+contents['bfr'],fill='#787878',font=self.fonts('思源黑体',26))
        #BFR图
        pic_bfr=contents['pic_bfr']
        pic_bfr=pic_bfr.resize((600,600*pic_bfr.size[1]//pic_bfr.size[0]))
        pic_bfr=pic_bfr.crop((0,int(600*pic_bfr.size[1]//pic_bfr.size[0]//3),pic_bfr.size[0],pic_bfr.size[1]))
        bg.paste(pic_bfr,(60,1140))
        #体适能雷达图
        _ico_radar=Image.open(os.path.join(self.material_dir,'UI图标','radar.png'))
        ico_radar=self.icon(_ico_radar,ico_size)
        bg.paste(ico_radar[0],(50,1400),mask=ico_radar[1])
        draw.text((120,1410),'体适能',fill='#787878',font=self.fonts('思源黑体',26))
        pic_radar=contents['pic_radar']
        pic_radar=pic_radar.resize((600,600*pic_radar.size[1]//pic_radar.size[0]))
        # pic_radar=pic_radar.crop((0,int(600*pic_radar.size[1]//pic_radar.size[0]//3),pic_radar.size[0],pic_radar.size[1]))
        bg.paste(pic_radar,(60,1460))

        #运动记录----------------------------------------------------
        bg_diary=Image.new('RGBA',(720,block_ht['b_diary']),color=colors['train_rec'])
        bg.paste(bg_diary,(0,y_diary))

        #训练日记标题
        _ico_dot=Image.open(os.path.join(self.material_dir,'UI图标','dot.png'))
        ico_dot=self.icon(_ico_dot,ico_size)
        y_diary_title=y_diary+40
        bg.paste(ico_dot[0],(50,y_diary_title),mask=ico_dot[1])
        draw.text((100,y_diary_title),'运动记录',fill='#787878',font=self.fonts('思源黑体',36))
        draw.line((50,y_diary_title+50,680,y_diary_title+50),fill='#787878')
        #有氧时长
        _ico_oxy=Image.open(os.path.join(self.material_dir,'UI图标','oxy_sport.png'))
        ico_oxy=self.icon(_ico_oxy,ico_size)
        bg.paste(ico_oxy[0],(100,y_diary_title+80),mask=ico_oxy[1])
        draw.text((160,y_diary_title+90),'有氧运动时长：'+contents['oxy_time'],fill='#787878',font=self.fonts('思源黑体',26))
        #总抗阻重量
        _ico_wt=Image.open(os.path.join(self.material_dir,'UI图标','dumbbell.png'))
        ico_wt=self.icon(_ico_wt,ico_size)
        bg.paste(ico_wt[0],(100,y_diary_title+145),mask=ico_wt[1])        
        draw.text((160,y_diary_title+150),'抗阻总重量：'+contents['muscle_wt'],fill='#787878',font=self.fonts('思源黑体',26))
        #各部位训练次数
        _ico_body=Image.open(os.path.join(self.material_dir,'UI图标','body.png'))
        ico_body=self.icon(_ico_body,ico_size)
        bg.paste(ico_body[0],(100,y_diary_title+205),mask=ico_body[1])   
        draw.text((160,y_diary_title+210),'各部位训练次数\n\n'+contents['each_part'],fill='#787878',font=self.fonts('思源黑体',26))

        #围度变化--------------------------------------------------------------------------------
        bg_msr=Image.new('RGBA',(720,block_ht['b_msr']),color=colors['msr_change'])
        bg.paste(bg_msr,(0,y_msr))
        y_msr_title=y_msr+40
        _ico_dot=Image.open(os.path.join(self.material_dir,'UI图标','dot.png'))
        ico_dot=self.icon(_ico_dot,ico_size)
        bg.paste(ico_dot[0],(50,y_msr_title),mask=ico_dot[1])
        draw.text((100,y_msr_title),'围度变化',fill='#787878',font=self.fonts('思源黑体',36))
        draw.line((50,y_msr_title+50,680,y_msr_title+50),fill='#787878')
        pic_msr_chart=contents['pic_msr_chart']
        pic_msr_chart=pic_msr_chart.resize((600,600*pic_msr_chart.size[1]//pic_msr_chart.size[0]))
        # pic_radar=pic_radar.crop((0,int(600*pic_radar.size[1]//pic_radar.size[0]//3),pic_radar.size[0],pic_radar.size[1]))
        bg.paste(pic_msr_chart,(60,y_msr_title+80))

        #饮食建议----------------------------------------------------------------------
        bg_diet=Image.new('RGBA',(720,block_ht['b_diet']),color=colors['diet'])
        bg.paste(bg_diet,(0,y_diet))
        y_diet_title=y_diet+40
        _ico_dot=Image.open(os.path.join(self.material_dir,'UI图标','dot.png'))
        ico_dot=self.icon(_ico_dot,ico_size)
        bg.paste(ico_dot[0],(50,y_diet_title),mask=ico_dot[1])
        draw.text((100,y_diet_title),'饮食建议',fill='#787878',font=self.fonts('思源黑体',36))
        draw.line((50,y_diet_title+50,680,y_diet_title+50),fill='#787878')
        composing.put_txt_img(draw=draw,tt=self.read_diet(),total_dis=diet_boxwid,xy=(70,y_diet_title+80),dis_line=diet_font_size*1.3,fill='#787878',font_name='思源黑体',font_size=diet_font_size,addSPC='yes',font_config_file=os.path.join(os.path.dirname(__file__),'configs','FontList.minghu.config'))

        #底部---------------------------------------------------
        bg_bottom=Image.new('RGBA',(720,block_ht['b_bottom']),color=colors['bottom'])
        bg.paste(bg_bottom,(0,y_bottom))
        qrcode=Image.open(os.path.join(self.ins_dir,ins+'二维码.jpg'))
        qrcode=qrcode.resize((100,100))
        bg.paste(qrcode,(50,y_bottom+25))

        slogan=Image.open(os.path.join(self.material_dir,'UI图标','slogan.png'))
        slogan=slogan.resize((330,slogan.size[1]*330//slogan.size[0]))
        m_slogan=slogan.split()[3]
        bg.paste(slogan,(225,y_bottom+30),mask=m_slogan)
 
        # bg.show()
        outimg=bg.convert('RGB')
        # outimg.show()
        outimg.save('C:\\Users\\jack\\Desktop\\demo0.jpg',quality=90,subsampling=0)

        print('完成')

if __name__=='__main__':
    #根据训练数据生成阶段报告
    p=PeroidSummary(place='seven')
    p.exp_chart(cus_name_input='SV001测试',ins='SVINS001周颖鑫',
                start_date='20210429',end_date='20210827',theme='lightgrey',
                ico_size=(40,40),diary_font_size=26,diet_font_size=26,diet_boxwid=580,logo_ht=72)
    # p.draw(cus='SV001测试',ins='SVINS001周颖鑫',start_time='20200115',end_time='20210820')
    # res=p.cal_data()
    # print(res)
    # # res['pic_bmi'].show()
    # res['pic_bfr'].show()

    #当天报告
    # p=FeedBackAfterClass(place='seven')
    # p.draw_new(cus='SV001测试',ins='SVINS001周颖鑫',date_input='20210817')
    # p.draw(cus='MH037廖程',ins='MHINS002韦越棋',date_input='20210824')
    # p.group_afterclass(ins='MHINS002韦越棋',date_input='20210727',open_dir='no')

    # 根据多次体测数据生成折线图
    # fitdata=FitData2Pic()
    # fitdata.to_pic(items=['chest','waist','hip'])

    #分组录入数据
    # p=GroupDataInput()
    # p.data_input()

    #计算体脂率
    # my=cals()
    # print(my.bfr(age=40,sex='男',ht=170,wt=63.8,waist=82,formula=1))


    # s='20211101'
    # e='20221222'
    # vsy=datetime.strptime(s,'%Y%m%d').year
    # vey=datetime.strptime(e,'%Y%m%d').year
    # vsm=datetime.strptime(s,'%Y%m%d').month
    # vem=datetime.strptime(e,'%Y%m%d').month
    # vsd=datetime.strptime(s,'%Y%m%d').day
    # ved=datetime.strptime(e,'%Y%m%d').day

    # delta_m=(vey-vsy)*12+(vem-vsm)

    # print(delta_m)    

