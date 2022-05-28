import os
import openpyxl
import re

def add_excel_col(col=2,fn='E:\\temp\\minghu\\会员\\会员资料\\MH015欧.xlsx'):
    try:
        wb = openpyxl.load_workbook(fn)
        ws = wb['训练情况']
        # 在第3列之前插入数据，这里序号是从1开始的
        ws.insert_cols(col)
        ws.insert_cols(col)
        # 插入数据
        ws['B1']='节次'
        ws['C1']='课程'
        ws.merge_cells('B1:B2')
        ws.merge_cells('C1:C2')
        wb.save(fn)
        print(fn,'……完成')
    except Exception as e:
        print(fn,'……',e)

def gp_add_col(col=2,dir='E:\\temp\\minghu\\会员\\会员资料'):
    for fn in os.listdir(dir):
        if re.match(r'MH\d{3}.*.xlsx',fn):
            add_excel_col(col=col,fn=os.path.join(dir,fn))




if __name__=='__main__':
    gp_add_col(col=2,dir='E:\\temp\\minghu\\会员\\会员资料')