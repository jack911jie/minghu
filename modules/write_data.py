import pandas as pd

# pd.set_option('max_columns', None) # 显示最多列数
pd.set_option('expand_frame_repr', False) # 当列太多时显示不清楚
pd.set_option('display.unicode.east_asian_width', True) #设置输出右对齐
import numpy as np
import openpyxl

class WriteData:
    def write_to_xlsx(self,input_dataframe,output_xlsx,sheet_name,cols=[],parse_date_col_name='时间',allow_dupliate='yes',dup_judge_by_cols=''):
        if parse_date_col_name:
            old=pd.read_excel(output_xlsx,sheet_name=sheet_name,parse_dates=[parse_date_col_name])
        else:
            old=pd.read_excel(output_xlsx,sheet_name=sheet_name)        

        if input_dataframe.shape[0]>0:        
            #不允许重复
            if allow_dupliate=='no': 
                input_dataframe=self.verify_data(df_old=old,df_new=input_dataframe,cols=cols,method='keep_new')
                #有判断重复列名的输入
                if dup_judge_by_cols:
                    book=openpyxl.load_workbook(output_xlsx)
                    writer=pd.ExcelWriter(output_xlsx,engine='openpyxl')
                    writer.book=book
                    writer.sheets=dict((ws.title,ws) for ws in book.worksheets)      
                    try:
                        log_txt=''
                        rows=[]
                        for itm in input_dataframe[dup_judge_by_cols].tolist():
                            row=old.index[old[dup_judge_by_cols]==itm].tolist()[-1]+1
                            rows.append(row)                        
                        for n,row in enumerate(rows):
                            df_input_row=input_dataframe.iloc[n].to_frame().T
                            df_input_row.to_excel(writer,sheet_name=sheet_name,startrow=row,index=False,header=False)  
                            log_txt=log_txt+str(input_dataframe.shape[0])+' 条数据追加完成，行号：'+str(row+1)+'\n'
                            log_txt=log_txt.strip()
                        writer.save()   
                    except Exception as err_cover:
                        writer.save()   
                        print('覆盖原记录出错：',err_cover)
                else:
                    err_not_cover='不允许重复时，请输入判断的列名（只允许一个列名）。'   
                    log_txt='无新记录，写入0条'                    
                    print('追加记录出错：',err_not_cover)
            #允许重复（直接追加）
            else:
                book=openpyxl.load_workbook(output_xlsx)
                writer=pd.ExcelWriter(output_xlsx,engine='openpyxl')
                writer.book=book
                writer.sheets=dict((ws.title,ws) for ws in book.worksheets)     
                old_rows=old.shape[0]       
                try: 
                    input_dataframe.to_excel(writer,sheet_name=sheet_name,startrow=old_rows+1,index=False,header=False)                  
                    log_txt=str(input_dataframe.shape[0])+' 条数据追加完成，行号：'+str(old_rows+2)+'-'+str(old_rows+input_dataframe.shape[0]+1)
                    writer.save()
                except Exception as err_not_cover:
                    writer.save()   
                    print('追加记录出错：',err_not_cover)  
                    log_txt='无新记录，写入0条'
                
        else:
            log_txt='无新记录，写入0条'

             

        return log_txt
    
    def verify_data(self,df_old,df_new,cols,method='keep_new'):
        if method=='keep_new':
            if len(cols)>0:
                df_diff=pd.concat([df_new,df_old,df_old]).drop_duplicates(subset=cols,keep=False,inplace=False)
            else:
                df_diff=pd.concat([df_new,df_old,df_old]).drop_duplicates(keep=False,inplace=False)
        elif method=='keep_old':
            if len(cols)>0:
                df_diff=pd.concat([df_new,df_old,df_new]).drop_duplicates(subset=cols,keep=False,inplace=False)
            else:
                df_diff=pd.concat([df_new,df_old,df_new]).drop_duplicates(keep=False,inplace=False)
        else:
            print('未设置方式')
            df_diff=pd.DataFrame()
        
        return df_diff

    def convert_column_to_date_format(self,file_path,sheet_name, column_name):
        try:
            wb = openpyxl.load_workbook(file_path,keep_vba=True)
            sheet = wb[sheet_name]

            # 获取列索引
            column_index = None
            for cell in sheet[1]:
                if cell.value == column_name:
                    column_index = cell.column_letter
                    break

            if column_index is not None:
                # 设置日期格式
                # date_style = NamedStyle(name='date_format')
                # date_style.number_format = numbers.FORMAT_DATE_XLSX14
                sheet.column_dimensions[column_index].width = 13  # 设置列宽
                for cell in sheet[column_index]:
                    cell.number_format='yyyy/mm/dd'
                wb.save(file_path)
            wb.close()
        except Exception as e:
            res=e
        res=''

        # print('{} 修改为短日期格式完成'.format(column_name))
        return res


if  __name__=='__main__':
    old='e:\\temp\\dddd.xlsx'
    dfin=pd.DataFrame(data={'日期':[20230310,20230625],'评价':['qqqdddd','sesfasfsa ']})
    p=WriteData()
    res=p.write_to_xlsx(input_dataframe=dfin,output_xlsx=old,sheet_name='Sheet1',cols=[],parse_date_col_name='',allow_dupliate='no',dup_judge_by_cols='日期')
    # allow_duplicate —— 允许重复，如设置为no，则dup_judge_by_cols必须输入
    # dup_judge_by_cols —— 当允许重复输入时，该参数为判断重复的列名标准，如：设置为表格中的日期项，则会覆盖相同日期的记录。
    print(res)